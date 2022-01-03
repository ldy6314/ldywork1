from urllib.parse import urlparse, urljoin
from flask import request, redirect, url_for, current_app, send_file
from uuid import uuid4
import os
from openpyxl import load_workbook
from extensions import db
from models import Subject, Class
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from io import BytesIO
from urllib.parse import quote


# 函数功能，传入当前url 跳转回当前url的前一个url
def redirect_back(backurl, **kwargs):
    for target in request.args.get('next'), request.referrer:
        if not target:
            continue
        if is_safe_url(target):
            return redirect(target)
    return redirect(url_for(backurl, **kwargs))


def is_safe_url(target):
    ref_url = urlparse(request.host_url)
    test_url = urlparse(urljoin(request.host_url, target))
    return test_url.scheme in ('http', 'https') and ref_url.netloc == test_url.netloc


def random_filename(filename):
    ext = os.path.splitext(filename)[1]
    new_filename = uuid4().hex + ext
    return new_filename


def get_subjects(filename):
    wb = load_workbook(filename)
    ws = wb.active
    minr = ws.min_row
    minc = ws.min_column
    maxr = ws.max_row
    maxc = ws.max_column

    subjects = []
    for row in ws.iter_rows(min_row=2, max_row=maxr, max_col=4):
        subject = []
        for cell in row:
            subject.append(cell.value)
        subjects.append(subject)
    return subjects


def to_class_id(class_string):
    grd = class_string[0]
    cls = class_string[1]
    grade_list = current_app.config['GRADE_LIST']
    class_of_grade = current_app.config['CLASS_OF_GRADE']
    try:
        grd_idx = grade_list.index(grd)
        cls_idx = int(cls)
    except IndexError:
        return -1
    return grd_idx*class_of_grade + cls_idx


def parser_class_id(class_id):
    grade_list = current_app.config['GRADE_LIST']
    class_of_grade = current_app.config['CLASS_OF_GRADE']
    grd = grade_list[class_id // class_of_grade]
    cls = class_id % class_of_grade
    return grd, cls


def get_time_list():
    time_list = db.session.query(Subject.time).all()
    time_list = [time[0] for time in set(time_list)]
    return time_list


def get_students_information(filename):
    print(filename)
    wb = load_workbook(filename)
    ws = wb.active
    grd = ws['B1'].value
    cls = ws['D1'].value
    infos = []
    for row in ws.iter_rows(min_row=3, max_row=50, max_col=7):
        info = []
        for cell in row:
            info.append(cell.value)
        if info[0]:
            infos.append(info)
    return grd, cls, infos


def get_users_information(filename):
    wb = load_workbook(filename)
    ws = wb.active
    infos = []
    for row in ws.iter_rows(min_row=2, max_row=50, max_col=5):
        info = []
        for cell in row:
            info.append(cell.value)
        if info[0]:
            infos.append(info)
    return infos


def get_class_info(class_id):
    cls = Class.query.filter_by(id=class_id).first()
    infos = []
    canceled_infos = []
    class_tot = 0
    subjects_cnt = 0
    students_cnt = len(cls.students)
    for student in cls.students:
        info = [student.id, student.name, student.contact1, student.contact2]
        tot = 0
        subjects = student.subjects
        canceled_list = []
        for i in range(4):
            try:
                if subjects[i].canceled:
                    canceled_list.append(subjects[i].name)
                info.append("{}<br>{}<br>{}".format(subjects[i].name, subjects[i].price, subjects[i].time))
                subjects_cnt += 1
                class_tot += subjects[i].price
                tot += subjects[i].price
            except IndexError:
                info.append("")
        info.append(tot)
        infos.append(info)
        if canceled_list:
            canceled_infos.append((student.id, student.name, ','.join(canceled_list)))

    tot_info = ("报名人数", students_cnt), ("报名人次", subjects_cnt), ("班级总费用", class_tot)
    return infos, tot_info, canceled_infos


def download_excel(filename, sheet_names, **kwargs):
    """
      col_name
      data_list
      head_merge_range
      border_range
      cols_width_info
    """
    wb = Workbook()
    for idx, sheet_name in enumerate(sheet_names):
        ws = wb.worksheets[idx]
        # 表头部分
        # 标题
        ws.append([sheet_name])
        # 每列的值
        col_name = kwargs.get('col_name', "")
        ws.append(col_name)
        # 数据部分
        data_list = kwargs.get('data_list', [])
        data = data_list[idx]
        for line in data:
            # print(i[0], i[1].name, i[1].time, i[1].price, i[1].remark)
            ws.append(line)

        # 格式部分
        head_merge_range = kwargs.get('head_merge_range', None)
        if head_merge_range:
            ws.merge_cells(head_merge_range)
        border = Border(
            left=Side(border_style='thin', color="FF000000"),
            right=Side(border_style='thin', color="FF000000"),
            bottom=Side(border_style='thin', color="FF000000"),
            top=Side(border_style='thin', color="FF000000")
        )
        align_center = Alignment(horizontal='center', vertical='center', wrapText=True)
        # 格式修饰区域1
        border_range = kwargs.get('border_range', None)
        if border_range:
            ws_area = ws[border_range]
            for row in ws_area:
                for cell in row:
                    cell.alignment = align_center
                    cell.border = border
        col_width_infos = kwargs.get('col_width_infos', {})
        for col in col_width_infos:
            ws.column_dimensions[col].width = col_width_infos[col]

        virtual_book = BytesIO()
        wb.save(virtual_book)
        virtual_book.seek(0)
    rv = send_file(virtual_book, as_attachment=True, attachment_filename=filename)
    rv.headers['Content-Disposition'] += ";filename*=utf-8' '{}".format(quote(filename))
    return rv
