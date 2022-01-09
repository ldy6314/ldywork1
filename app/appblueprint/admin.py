from flask import Blueprint, send_file, session, jsonify, flash, current_app, redirect, url_for
from forms import AddForm, AddSubjectForm, UploadClassForm, AddStudentForm, UploadSubjectsForm, EditSubjectForm, \
    UploadUserForm, EditStudentForm, AddFstudentForm
from flask import render_template
from models import Subject, Student, User, Class, Fstudent
from extensions import db
from flask_login import current_user
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from io import BytesIO
from flask_login import login_required
from utils import redirect_back, random_filename, get_subjects, get_time_list, parser_class_id, \
    get_students_information, to_class_id, get_users_information, get_class_info, download_excel
import os
from openpyxl.worksheet.datavalidation import DataValidation
from urllib.parse import quote

admin_bp = Blueprint('admin', __name__)


@admin_bp.before_request
@login_required
def login_required():
    pass


@admin_bp.route('/')
def index():
    return "I am appblueprint index"


@admin_bp.route('/add', methods=['GET', 'POST'])
def add():
    form = AddForm()
    if form.validate_on_submit():
        for i in form.data:
            print(i, form.data[i])

    return render_template('adduser.html', form=form)


@admin_bp.route('/add_student', methods=['GET', 'POST'])
def add_student():
    # return render_template("over.html")
    form1 = AddStudentForm()
    time_list = get_time_list()
    sublist = [form1.sub1, form1.sub2, form1.sub3, form1.sub4]
    for item in zip(sublist, time_list):
        item[0].label.text = item[1] + "课程"
        item[0].choices = [subject.name for subject in Subject.query.filter_by(time=item[1]).all()]
        item[0].choices.insert(0, "不选")
    form1 = AddStudentForm()
    time_list = get_time_list()
    sublist = [form1.sub1, form1.sub2, form1.sub3, form1.sub4]
    for item in zip(sublist, time_list):
        item[0].label.text = item[1] + "课程"
        item[0].choices = [subject.name for subject in Subject.query.filter_by(time=item[1]).all()]
        item[0].choices.insert(0, "不选")

    if form1.validate_on_submit():
        print("{}{}添加学生".format(*parser_class_id(current_user.class_id)))
        class_id = current_user.class_id
        data = form1.data
        name = data['name']
        con1 = data['contact1']
        con2 = data['contact2']
        subs = [data['sub1'], data['sub2'], data['sub3'], data['sub4']]
        res = Student.query.filter_by(name=name, class_id=class_id).all()
        if not res:
            student = Student(name=name, class_id=class_id, contact1=con1, contact2=con2)

            for sub in subs:
                if sub != "不选":
                    sub = Subject.query.filter_by(name=sub).first()
                    if sub:
                        student.subjects.append(sub)

            if len(student.subjects):
                db.session.add(student)
                db.session.commit()
                flash('添加成功')
            else:
                flash('至少选择一项科目')

        else:
            flash('该学生已经存在')

    return render_template("addstudent.html", form1=form1)


@admin_bp.route('/add_subject', methods=['GET', 'POST'])
def add_subject():
    form = AddSubjectForm()
    if form.validate_on_submit():
        subject = Subject(name=form.data['name'], time=form.data['time'], price=form.data['price'],
                          remark=form.data['remark'])

        db.session.add(subject)
        db.session.commit()
        # res = Subject.query.all()
        # for i in res:
        #     print(i.name)
        return jsonify({'result': "插入成功", "type": "alert alert-success"})
    else:
        return jsonify({'result': "项目已经存在无法添加", "type": "alert  alert-danger"})


@admin_bp.route('/download_subjects')
def download_subjects():
    filename = "课程信息表.xlsx"
    sheet_names = ["课程信息表"]
    col_name = ["编号", "名称", "时间", "价格", "备注"]
    head_merge_range = "A1:E1"
    col_width_infos = {"B": 30, "C": 30, "E": 30}
    data_list = [[]]
    res = Subject.query.all()
    for idx, subject in enumerate(res, 1):
        info = [idx, subject.name, subject.time, subject.price, subject.remark]
        data_list[0].append(info)
    border_range = ["A{}:E{}".format(1, len(res) + 2)]
    return download_excel(filename, sheet_names,
                          col_name=col_name,
                          data_list=data_list,
                          head_merge_range=head_merge_range,
                          border_range=border_range,
                          col_width_infos=col_width_infos)


@admin_bp.route('/school_admin', methods=['GET', 'POST'])
def school_admin():
    permission = current_user.permission
    if permission == 2:
        res1 = Fstudent.query.all()
        finfos = []
        for i in res1:
            grd, cls = parser_class_id(i.class_id)
            line = [i.name, grd + str(cls)]
            subjects = i.subjects
            for idx in range(4):
                if idx < len(subjects):
                    info = subjects[idx].name + '<br>' + subjects[idx].time
                else:
                    info = ""
                line.append(info)
            line.append(i.id)
            finfos.append(line)
        form = UploadSubjectsForm()
        subjects = Subject.query.all()
        form1 = AddSubjectForm()
        time_list = db.session.query(Subject.time).all()
        time_list = [time[0] for time in set(time_list)]
        form1.time.choices = time_list
        form2 = AddForm()
        users = User.query.all()
        subjects_info = []
        subject_list = Subject.query.all()
        for subject in subject_list:
            subjects_info.append((subject.name, len(subject.students), subject.id))
        class_list = Class.query.all()
        class_infos = []
        student_cnt = 0
        subject_cnt = 0
        tot_cost = 0
        for cls in class_list:
            info = get_class_info(cls.id)[1]
            clsname = parser_class_id(cls.id)
            clsname = clsname[0] + str(clsname[1])
            class_infos.append((clsname, info[0][1], info[1][1], info[2][1]))
            student_cnt += info[0][1]
            subject_cnt += info[1][1]
            tot_cost += info[2][1]

        tot_info = "合计", student_cnt, subject_cnt, tot_cost
        class_infos.append(tot_info)
        return render_template('schooladmin.html', form=form, form1=form1, form2=form2, subjects=subjects, users=users,
                               subjects_info=subjects_info, class_infos=class_infos, to_class_id=to_class_id,
                               finfos=finfos)
    else:
        return render_template("permission_deny.html")


@admin_bp.route('/class_admin')
def class_admin():
    form = UploadClassForm()
    permission = current_user.permission
    class_id = current_user.class_id
    if permission == 1:
        infos, tot_info, canceled_infos = get_class_info(class_id)
        return render_template('classadmin.html', form=form, infos=infos, tot_info=tot_info,
                               canceled_infos=canceled_infos)
    return render_template("permission_deny.html")


@admin_bp.route('/edit_subject/<int:subject_id>', methods=['GET', 'POST'])
def edit_subject(subject_id):
    time_list = get_time_list()
    form = EditSubjectForm()
    form.time.choices = time_list
    subject = Subject.query.get(subject_id)
    if form.validate_on_submit():
        subject.time = form.time.data
        subject.price = form.price.data
        subject.remark = form.remark.data
        subject.canceled = 1 if form.canceled.data else 0
        if subject.canceled:
            subject.name = subject.name + "(已取消)"
        else:
            if "(已取消)" in subject.name:
                subject.name = subject.name.strip("(已取消)")
        db.session.commit()
        flash("课程修改成功")
        return redirect(url_for('admin.school_admin'))

    name = subject.name
    tm = subject.time
    pr = subject.price
    rm = subject.remark
    cn = True if subject.canceled == 1 else False
    form.remark.data = rm
    form.price.data = pr
    form.time.choices = time_list
    form.time.data = tm
    form.canceled.data = cn
    return render_template('editsubject.html', form=form, subject_name=name)


@admin_bp.route('/delete/<int:subject_id>', methods=['GET', 'POST'])
def delete_subject(subject_id):
    subject = Subject.query.filter_by(id=subject_id).first()
    db.session.delete(subject)
    flash("已经删除 {}".format(subject.name))
    db.session.commit()
    return redirect_back('/')


@admin_bp.route('/upload_subjects', methods=['GET', 'POST'])
def upload_subjects():
    path = os.path.join(current_app.root_path, 'upload\\')
    form = UploadSubjectsForm()
    if form.validate_on_submit():
        f = form.file.data
        filename = random_filename(f.filename)
        f.save(path + filename)
        try:
            subjects = get_subjects(path + filename)
        except Exception as e:
            print(e)
            flash('上传的表格不正确')
            return redirect_back('/')
        infos = []
        for i in subjects:
            if not i[0]:
                break
            res = Subject.query.filter_by(name=i[0]).first()
            if not res:
                subject = Subject(name=i[0], time=i[1], price=str(i[2]).strip('元'), canceled=0, remark="")
                infos.append(subject)
            else:
                flash('{}已经存在！'.format(i[0]))

        for i in infos:
            try:
                db.session.add(i)
                db.session.commit()
            except Exception as e:
                print(e)

    else:
        flash('上传失败')
    return redirect_back('/')


@admin_bp.route('/download_class_table')
def download_class_table():
    class_id = current_user.class_id
    grd, cls = parser_class_id(class_id)
    # print(grd, cls)
    time_list = get_time_list()
    subject_list = []
    for time in time_list:
        subjects = db.session.query(Subject.name).filter_by(time=time).all()
        subjects = [i[0] for i in subjects]
        subject_list.append(','.join(subjects))
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.append(["年级", "", "班级", ""])
    col_name = ["姓名", "联系电话1", "联系电话2"]
    tm_lst = [tm + '项目' for tm in time_list]
    col_name.extend(tm_lst)
    ws.append(col_name)
    dv = DataValidation(type="list", formula1='"一,二,三,四,五,六"', allow_blank=True)
    dv.error = "输入的值不在下拉列表中"
    dv.errorTitle = "错误"
    dv.prompt = "从下拉列表中选择年级"
    dv.promptTitle = "列表选择"
    dv.add('B1')
    ws.add_data_validation(dv)
    dv1 = DataValidation(type="list", formula1='"1,2,3,4,5,6,7,8"', allow_blank=True)
    dv1.error = "输入的值不在下拉列表中"
    dv1.errorTitle = "错误"
    dv1.prompt = "从下拉列表中选择班级"
    dv1.promptTitle = "列表选择"
    dv1.add('D1')
    ws.add_data_validation(dv1)
    ws['B1'].value = grd
    ws['D1'].value = cls
    for infos in zip(subject_list, ["D3:D50", "E3:E50", "F3:F50", "G3:G50"]):
        dv = DataValidation(type="list", formula1='"{}"'.format(infos[0]), allow_blank=True)
        dv.error = "输入的值不在下拉列表中"
        dv.errorTitle = "错误"
        dv.prompt = "请从下拉列表中选择课程"
        dv.promptTitle = "列表选择"
        dv.add(infos[1])
        ws.add_data_validation(dv)

    border = Border(
        left=Side(border_style='thin', color="FF000000"),
        right=Side(border_style='thin', color="FF000000"),
        bottom=Side(border_style='thin', color="FF000000"),
        top=Side(border_style='thin', color="FF000000")
    )
    align_center = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws_area = ws["A1:G50"]
    for row in ws_area:
        for cell in row:
            cell.alignment = align_center
            cell.border = border
    for i in "BCDEFG":
        ws.column_dimensions[i].width = 18

    virtual_book = BytesIO()
    wb.save(virtual_book)
    virtual_book.seek(0)
    rv = send_file(virtual_book, as_attachment=True, attachment_filename="class_table.xlsx")
    rv.headers['Content-Disposition'] += ";filename*=utf-8' '{}{}{}.xlsx".format(quote(grd), cls, quote("班级俱乐部报名表"))
    return rv


@admin_bp.route("/upload_class_table", methods=['GET', 'POST'])
def upload_class_table():
    return render_template("over.html")
    path = os.path.join(current_app.root_path, 'upload\\')
    # print("path=", path)
    form = UploadSubjectsForm()
    if form.validate_on_submit():
        print("{}{}班上传表格".format(*parser_class_id(current_user.class_id)))
        f = form.file.data
        filename = f.filename
        f.save(path + filename)
        try:
            grd, cls, infos = get_students_information(path + filename)
            class_id = to_class_id(grd + str(cls))
            if class_id != current_user.class_id:
                flash('班级填写不正确,请修改')
                return redirect_back('rootbp.index')

            for i in infos:
                res = Student.query.filter_by(class_id=class_id, name=i[0]).all()
                if not res:
                    student = Student(class_id=class_id, name=i[0], contact1=i[1] if i[1] else "", contact2=i[2] if i[2]
                    else "")

                    for sub in i[-4:]:
                        if sub:
                            sub = Subject.query.filter_by(name=sub).first()
                            if not sub:
                                flash('表格填写有错误')
                                return redirect_back('admin.class_admin')
                            student.subjects.append(sub)
                else:
                    flash("学生 {} 已经存在".format(i[0]))
                db.session.commit()
            flash("上传成功")
        except Exception as e:
            print(e)
            flash('上传的表格不正确')
            return redirect_back('admin.class_admin')

    else:
        flash('上传失败')
    return redirect_back('admin.class_admin')


@admin_bp.route('/upload_user_table', methods=['GET', 'POST'])
def upload_users_table():
    form = UploadUserForm()
    path = os.path.join(current_app.root_path, 'upload\\')

    if form.validate_on_submit():
        f = form.file.data
        filename = f.filename
        f.save(path + filename)
        try:
            infos = get_users_information(path + filename)
            for i in infos:
                res = User.query.filter_by(username=i[0]).all()
                if not res:
                    name = "{}{}班主任".format(i[2], i[3])
                    class_id = to_class_id(i[2] + str(i[3]))
                    print(name, class_id)
                    user = User(username=i[0], class_id=class_id, name=name, permission=i[4])
                    user.set_password(i[0])
                    db.session.add(user)
                    print(i[0], "添加成功")
                else:
                    flash("账号 {} 已经存在".format(i[0]))
                db.session.commit()
            flash("上传成功")
        except Exception as e:
            print(e)
            flash('上传的表格不正确')
            return redirect_back('/')

        else:
            flash('上传失败')
            return redirect_back('/')

    else:
        return render_template("upload_user_table.html", form=form)


@admin_bp.route('/edit_student/<int:student_id>', methods=['GET', 'POST'])
def edit_student(student_id):
    form1 = EditStudentForm()
    form1.submit.label.text = "修改"
    time_list = get_time_list()
    sublist = [form1.sub1, form1.sub2, form1.sub3, form1.sub4]
    for item in zip(sublist, time_list):
        item[0].label.text = item[1] + "课程"
        item[0].choices = [subject.name for subject in Subject.query.filter_by(time=item[1]).all()]
        item[0].choices.insert(0, "不选")
    student = Student.query.filter_by(id=student_id).first()
    if form1.validate_on_submit():
        grd, cls = parser_class_id(current_user.class_id)
        name = student.name
        print("{}{}修改学生{}信息".format(grd, cls, student.name))
        data = form1.data
        con1 = data['contact1']
        con2 = data['contact2']
        subs = [data['sub1'], data['sub2'], data['sub3'], data['sub4']]
        num_of_subs = 4 - subs.count('不选')
        if not num_of_subs:
            db.session.delete(student)
            db.session.commit()
            flash('因修改该后学生 {} 没有任何科目，删除学生 {}'.format(name, name))
            return redirect(url_for("admin.class_admin"))
        else:
            student.contact1 = con1
            student.contact2 = con2
            student.subjects.clear()
            for sub in subs:
                if sub != "不选":
                    sub = Subject.query.filter_by(name=sub).first()
                    if sub:
                        student.subjects.append(sub)
            db.session.commit()
            flash('修改成功')

    name = student.name
    form1.contact1.data = student.contact1
    form1.contact2.data = student.contact2
    subjects = student.subjects
    for subject in subjects:
        idx = time_list.index(subject.time)
        sublist[idx].data = subject.name

    return render_template("editstudent.html", form=form1, name=name)


@admin_bp.route('/class_info/<int:class_id>')
def \
        show_class_info(class_id):
    infos, tot_info, _ = get_class_info(class_id)
    grd, cls = parser_class_id(class_id)
    class_name = grd + str(cls)
    return render_template('class_info.html', infos=infos, tot_info=tot_info, class_id=class_id,
                           class_name=class_name)


@admin_bp.route('/subject_info/<int:subject_id>')
def show_subject_info(subject_id):
    res = Subject.query.filter_by(id=subject_id).first()
    students = res.students
    fstudents = res.fstudents
    subject_name = res.name
    students.sort(key=lambda x: x.class_id)
    infos = []
    for idx, student in enumerate(students, 1):
        grd, cls = parser_class_id(student.class_id)
        infos.append((idx, grd + str(cls), student.name))
    for idx, fstudent in enumerate(fstudents, len(students) + 1):
        grd, cls = parser_class_id(fstudent.class_id)
        infos.append((idx, grd + str(cls), fstudent.name))
    return render_template('subject_info.html', infos=infos, subject_id=subject_id, subject_name=subject_name)


@admin_bp.route('/download_class_info/<int:class_id>')
def download_class_info(class_id):
    infos, tot_info, _ = get_class_info(class_id)
    grd, cls = parser_class_id(class_id)
    class_name = grd + str(cls)
    filename = "{}报名信息表.xlsx".format(class_name)
    sheet_names = [filename.strip('.xlsx')]
    col_name = ["姓名", "电话1", "电话2", "项目1", "项目2", "项目3", "项目4", "合计"]
    head_merge_range = "A1:H1"
    col_width_infos = {"B": 15, "C": 15, "D": 25, "E": 25, "F": 25, "G": 25}
    data_list = [[]]
    for student in infos:
        info = [student[1], student[2], student[3]]
        for i in range(4, 8):
            info.append(student[i].replace('<br>', "\n"))
        info.append(student[8])
        data_list[0].append(info)
    line = []
    for info in tot_info:
        line.extend(info)
    data_list[0].append(line)
    border_range = ["A{}:H{}".format(1, len(infos) + 3)]
    return download_excel(filename, sheet_names,
                          col_name=col_name,
                          data_list=data_list,
                          head_merge_range=head_merge_range,
                          border_range=border_range,
                          col_width_infos=col_width_infos)


@admin_bp.route('/download_subject_table/<int:subject_id>')
def download_subject_table(subject_id):
    res = Subject.query.filter_by(id=subject_id).first()
    subject_name = res.name
    students = res.students
    fstudents = res.fstudents
    students.sort(key=lambda s: s.class_id)
    filename = "{}俱乐部人员名单.xlsx".format(subject_name)
    sheet_names = [filename.strip('.xlsx')]
    col_name = ["编号", "班级", "姓名"]
    head_merge_range = "A1:C1"
    col_width_infos = {"B": 15, "C": 15}
    data_list = [[]]
    for idx, student in enumerate(students, 1):
        grd, cls = parser_class_id(student.class_id)
        info = [idx, grd + str(cls), student.name]
        data_list[0].append(info)

    for idx, fstudent in enumerate(fstudents, len(students) + 1):
        grd, cls = parser_class_id(fstudent.class_id)
        info = [idx, grd + str(cls), fstudent.name]
        data_list[0].append(info)

    border_range = ["A{}:C{}".format(2, len(students) + len(fstudents) + 2)]
    return download_excel(filename, sheet_names,
                          col_name=col_name,
                          data_list=data_list,
                          head_merge_range=head_merge_range,
                          border_range=border_range,
                          col_width_infos=col_width_infos)


@admin_bp.route("/add_fstudent", methods=['GET', 'POST'])
def add_fstudent():
    form1 = AddFstudentForm()
    time_list = get_time_list()
    sublist = [form1.sub1, form1.sub2, form1.sub3, form1.sub4]
    for item in zip(sublist, time_list):
        item[0].label.text = item[1] + "课程"
        item[0].choices = [subject.name for subject in Subject.query.filter_by(time=item[1]).all()]
        item[0].choices.insert(0, "不选")

    if form1.validate_on_submit():
        data = form1.data
        grd = data['grd']
        cls = data['cls']
        class_id = to_class_id(grd + cls)
        # print(class_id)
        name = data['name']
        con1 = data['contact1']
        con2 = data['contact2']
        subs = [data['sub1'], data['sub2'], data['sub3'], data['sub4']]
        # print(grd, cls, name, con1, con2, subs, class_id)
        res = Fstudent.query.filter_by(name=name, class_id=class_id).first()
        if not res:
            fstudent = Fstudent(name=name, class_id=class_id, contact1=con1, contact2=con2)
            for sub in subs:
                if sub != "不选":
                    sub = Subject.query.filter_by(name=sub).first()
                    if sub:
                        fstudent.subjects.append(sub)

            if len(fstudent.subjects):
                db.session.add(fstudent)
                db.session.commit()
                flash('添加成功')
            else:
                flash('至少选择一项科目')

        else:
            flash('该学生已经存在')

    return render_template("add_fstudent.html", form=form1)


@admin_bp.route("/edit_fstudent/<int:fstudent_id>", methods=["POST", 'GET'])
def edit_fstudent(fstudent_id):
    form1 = EditStudentForm()
    form1.submit.label.text = "修改"
    time_list = get_time_list()
    sublist = [form1.sub1, form1.sub2, form1.sub3, form1.sub4]
    for item in zip(sublist, time_list):
        item[0].label.text = item[1] + "课程"
        item[0].choices = [subject.name for subject in Subject.query.filter_by(time=item[1]).all()]
        item[0].choices.insert(0, "不选")
    fstudent = Fstudent.query.filter_by(id=fstudent_id).first()
    if form1.validate_on_submit():
        grd, cls = parser_class_id(fstudent.class_id)
        name = fstudent.name
        print("{}{}修改学生{}信息".format(grd, cls, fstudent.name))
        data = form1.data
        con1 = data['contact1']
        con2 = data['contact2']
        subs = [data['sub1'], data['sub2'], data['sub3'], data['sub4']]
        num_of_subs = 4 - subs.count('不选')
        if not num_of_subs:
            db.session.delete(fstudent)
            db.session.commit()
            flash('因修改该后学生 {} 没有任何科目，删除学生 {}'.format(name, name))
            return redirect(url_for("admin.school_admin"))
        else:
            fstudent.contact1 = con1
            fstudent.contact2 = con2
            fstudent.subjects.clear()
            for sub in subs:
                if sub != "不选":
                    sub = Subject.query.filter_by(name=sub).first()
                    if sub:
                        fstudent.subjects.append(sub)
            db.session.commit()
            flash('修改成功')

    name = fstudent.name
    form1.contact1.data = fstudent.contact1
    form1.contact2.data = fstudent.contact2
    subjects = fstudent.subjects
    for subject in subjects:
        idx = time_list.index(subject.time)
        sublist[idx].data = subject.name

    return render_template("editstudent.html", form=form1, name=name)


@admin_bp.route("/delete_fstudent/<int:fstudent_id>", methods=["POST", 'GET'])
def delete_fstudent(fstudent_id):
    fstudent = Fstudent.query.filter_by(id=fstudent_id).first()
    db.session.delete(fstudent)
    flash("已经删除 {}".format(fstudent.name))
    db.session.commit()
    return redirect_back('admin.school_admin')

@admin_bp.route('/upload_fstudents_table', methods=['GET', 'POST'])
def upload_fstudents_table():
    path = os.path.join(current_app.root_path, 'upload\\')
    # print("path=", path)
    form = UploadSubjectsForm()
    if form.validate_on_submit():
        f = form.file.data
        filename = f.filename
        f.save(path + filename)
        try:
            grd, cls, infos = get_students_information(path + filename, max_col=9, mode=1)
            for i in infos:
                grd, cls = i[1], i[2]
                class_id = to_class_id(grd + str(cls))
                res = Fstudent.query.filter_by(class_id=class_id, name=i[0]).all()
                if not res:
                    fstudent = Fstudent(class_id=class_id, name=i[0], contact1=i[3] if i[3] else "", contact2=i[4] if
                    i[4] else "")

                    for sub in i[-4:]:
                        if sub:
                            sub = Subject.query.filter_by(name=sub).first()
                            if not sub:
                                flash('表格填写有错误')
                                return redirect_back('admin.class_admin')
                            fstudent.subjects.append(sub)
                else:
                    flash("学生 {} 已经存在".format(i[0]))
                db.session.commit()
            flash("上传成功")
        except Exception as e:
            print(e)
            flash('上传的表格不正确')
            return redirect_back('admin.class_admin')

    return render_template('upload_file.html', form=form)


@admin_bp.route('/download_empty_table')
def download_empty_table():
    time_list = get_time_list()
    subject_list = []
    for time in time_list:
        subjects = db.session.query(Subject.name).filter_by(time=time).all()
        subjects = [i[0] for i in subjects]
        subject_list.append(','.join(subjects))
    wb = Workbook()
    ws = wb.worksheets[0]
    border = Border(
        left=Side(border_style='thin', color="FF000000"),
        right=Side(border_style='thin', color="FF000000"),
        bottom=Side(border_style='thin', color="FF000000"),
        top=Side(border_style='thin', color="FF000000")
    )
    align_center = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws.append(["教师子女报名表"])
    ws.merge_cells("A1:I1")
    ws["A1"].alignment = align_center
    col_name = ["姓名", "年级", "班级", "联系电话1", "联系电话2"]
    tm_lst = [tm + '项目' for tm in time_list]
    col_name.extend(tm_lst)
    ws.append(col_name)
    dv = DataValidation(type="list", formula1='"一,二,三,四,五,六"', allow_blank=True)
    dv.error = "输入的值不在下拉列表中"
    dv.errorTitle = "错误"
    dv.prompt = "从下拉列表中选择年级"
    dv.promptTitle = "列表选择"
    dv.add('B3:B50')
    ws.add_data_validation(dv)
    dv1 = DataValidation(type="list", formula1='"1,2,3,4,5,6,7,8"', allow_blank=True)
    dv1.error = "输入的值不在下拉列表中"
    dv1.errorTitle = "错误"
    dv1.prompt = "从下拉列表中选择班级"
    dv1.promptTitle = "列表选择"
    dv1.add('C3:C50')
    ws.add_data_validation(dv1)
    for infos in zip(subject_list, ["F3:F50", "G3:G50", "H3:H50", "I3:I50"]):
        dv = DataValidation(type="list", formula1='"{}"'.format(infos[0]), allow_blank=True)
        dv.error = "输入的值不在下拉列表中"
        dv.errorTitle = "错误"
        dv.prompt = "请从下拉列表中选择课程"
        dv.promptTitle = "列表选择"
        dv.add(infos[1])
        ws.add_data_validation(dv)

    ws_area = ws["A2:I50"]
    for row in ws_area:
        for cell in row:
            cell.alignment = align_center
            cell.border = border
    for i in "DEFGHI":
        ws.column_dimensions[i].width = 18

    virtual_book = BytesIO()
    wb.save(virtual_book)
    virtual_book.seek(0)
    rv = send_file(virtual_book, as_attachment=True, attachment_filename="class_table.xlsx")
    rv.headers['Content-Disposition'] += ";filename*=utf-8' '{}.xlsx".format(quote("教师子女报名表"))
    return rv


@admin_bp.route('/download_all_subjects')
def download_all_subjects():
    data_list = []
    sheet_names = []
    filename = "{}.xlsx".format("俱乐部名单汇总表")
    border_range = []
    res = Subject.query.all()
    for subject in res:
        data_list.append([])
        subject_name = subject.name
        students = subject.students
        fstudents = subject.fstudents
        students.sort(key=lambda s: s.class_id)

        sheet_names.append(subject_name+'俱乐部名单')
        col_name = ["编号", "班级", "姓名", "电话1", '电话2']
        head_merge_range = "A1:E1"
        col_width_infos = {"B": 15, "C": 15, "D": 20, "E": 20 }
        for idx, student in enumerate(students, 1):
            grd, cls = parser_class_id(student.class_id)
            con1 = student.contact1 if student.contact1 else ""
            con2 = student.contact2 if student.contact2 else ""
            info = [idx, grd + str(cls), student.name, con1, con2]
            data_list[-1].append(info)

        for idx, fstudent in enumerate(fstudents, len(students) + 1):
            grd, cls = parser_class_id(fstudent.class_id)
            con1 = fstudent.contact1 if student.contact1 else ""
            con2 = fstudent.contact2 if student.contact2 else ""
            info = [idx, grd + str(cls), fstudent.name, con1, con2]
            data_list[-1].append(info)

        border_range.append("A{}:E{}".format(2, len(students) + len(fstudents) + 2))
    return download_excel(filename, sheet_names,
                          col_name=col_name,
                          data_list=data_list,
                          head_merge_range=head_merge_range,
                          border_range=border_range,
                          col_width_infos=col_width_infos)


@admin_bp.route('/download_all_classes')
def download_all_classes():
    data_list = []
    sheet_names = []
    filename = "{}.xlsx".format("全校班级总表")
    border_range = []
    res = Class.query.all()
    for cls_ in res:
        data_list.append([])
        class_id = cls_.id
        infos, tot_info, _ = get_class_info(class_id)
        grd, cls = parser_class_id(class_id)
        class_name = grd + str(cls)
        sheet_names.append(class_name+'报名汇总表')
        col_name = ["姓名", "电话1", "电话2", "项目1", "项目2", "项目3", "项目4", "合计"]
        head_merge_range = "A1:H1"
        col_width_infos = {"B": 15, "C": 15, "D": 25, "E": 25, "F": 25, "G": 25}

        for student in infos:
            info = [student[1], student[2], student[3]]
            for i in range(4, 8):
                info.append(student[i].replace('<br>', "\n"))
            info.append(student[8])
            data_list[-1].append(info)
        line = []
        for info in tot_info:
            line.extend(info)
        data_list[-1].append(line)
        border_range.append("A{}:H{}".format(1, len(infos) + 3))
    return download_excel(filename, sheet_names,
                          col_name=col_name,
                          data_list=data_list,
                          head_merge_range=head_merge_range,
                          border_range=border_range,
                          col_width_infos=col_width_infos)


@admin_bp.route('/download_school_infos')
def download_school_infos():
    class_list = Class.query.all()
    class_infos = []
    student_cnt = 0
    subject_cnt = 0
    tot_cost = 0
    for cls in class_list:
        info = get_class_info(cls.id)[1]
        clsname = parser_class_id(cls.id)
        clsname = clsname[0] + str(clsname[1])
        class_infos.append((clsname, info[0][1], info[1][1], info[2][1]))
        student_cnt += info[0][1]
        subject_cnt += info[1][1]
        tot_cost += info[2][1]

    tot_info = "合计", student_cnt, subject_cnt, tot_cost
    class_infos.append(tot_info)
    data_list = [class_infos]
    filename = "班级汇总表.xlsx"
    sheet_names = ['全校班级汇总表']
    col_name = ['班级', '报名人数', '报名人次', '收费合计']
    border_range = ["A{}:D{}".format(2, len(class_infos)+2)]
    head_merge_range = "A1:D1"
    col_width_infos = []
    return download_excel(filename, sheet_names,
                          col_name=col_name,
                          data_list=data_list,
                          head_merge_range=head_merge_range,
                          border_range=border_range,
                          col_width_infos=col_width_infos)
