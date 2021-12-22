from openpyxl import load_workbook


def get_subjects(filename):
    wb = load_workbook(filename)
    ws = wb.active

    minr = ws.min_row
    minc = ws.min_column
    maxr = ws.max_row
    maxc = ws.max_column

    subjects = []
    for row in ws.iter_rows(min_row=2, max_row=maxr, max_col=maxc-1):
        subject = []
        for cell in row:
            subject.append(cell.value)
        subjects.append(subject)
    return subjects
